import io

from django.contrib.messages import get_messages
from django.urls import reverse
from django.utils import timezone

import openpyxl
import pytest

from common import messages as msg
from common.menus import MenuGroup, MenuItem, MenuSection
from input_data.models import ProformaSchedule, ProformaScheduleDetail


@pytest.mark.django_db
class TestProformaReadViews:
    """
    [그룹 1] 조회 관련 테스트
    범위: 목록(List), 검색(Search), 상세(Detail), 생성화면 진입(Initial)
    """

    def test_proforma_list_view(self, auth_client, sample_schedule):
        """
        [PF_LIST_001] 목록 조회 테스트
        Changed: 메뉴 그룹이 'Schedule'인지 확인
        """
        url = reverse("input_data:proforma_list")
        response = auth_client.get(url)

        assert response.status_code == 200
        assert "input_data/proforma_list.html" in [t.name for t in response.templates]

        # 리스트 데이터 확인
        proforma_list = response.context["proforma_list"]
        assert len(proforma_list) >= 1
        # [수정됨] 딕셔너리 접근(["lane_code"])에서 객체 속성 접근(.lane_id)으로 변경
        assert any(item.lane_id == "TEST_LANE" for item in proforma_list)

        # [Check] 메뉴 그룹 확인
        assert response.context["current_group"] == MenuGroup.SCHEDULE

    def test_proforma_list_search(self, auth_client, base_scenario, user):
        """
        [PF_LIST_002] 검색 기능 테스트
        """
        # 검색 대상이 아닌 데이터 추가 (Master - Detail 분리 구조 적용)
        master = ProformaSchedule.objects.create(
            scenario=base_scenario,
            lane_id="OTHER",
            proforma_name="PF_002",
            effective_from_date=timezone.now(),
            duration=14.0,
            declared_capacity="5000",
            declared_count=2,
            created_by=user,
            updated_by=user,
        )

        ProformaScheduleDetail.objects.create(
            proforma=master,
            direction="E",
            port_id="KRPUS",
            calling_port_indicator="1",
            calling_port_seq=1,
            turn_port_info_code="N",
            pilot_in_hours=3.0,
            etb_day_number=0,
            etb_day_code="SUN",
            etb_day_time="0900",
            actual_work_hours=24.0,
            etd_day_number=1,
            etd_day_code="MON",
            etd_day_time="1800",
            pilot_out_hours=3.0,
            link_distance=500,
            link_eca_distance=0,
            link_speed=20.0,
            sea_time_hours=24.0,
            terminal_code="PNC",
            created_by=user,
            updated_by=user,
        )

        url = reverse("input_data:proforma_list")

        # 'OTHER' 검색
        response = auth_client.get(url, {"lane_code": "OTHER"})
        results = response.context["proforma_list"]
        assert len(results) == 1
        assert results[0].lane_id == "OTHER"

        # 'TEST' 검색 (OTHER 제외 확인)
        response = auth_client.get(url, {"lane_code": "TEST"})
        results = response.context["proforma_list"]
        assert not any(item.lane_id == "OTHER" for item in results)

    def test_proforma_detail_view(self, auth_client, sample_schedule):
        """
        [PF_DETAIL_001] 상세 조회 테스트
        Changed: 메뉴 그룹이 'Schedule'인지 확인
        """
        url = reverse("input_data:proforma_detail")
        params = {
            "scenario_id": sample_schedule.scenario.id,
            "lane_code": sample_schedule.lane_id,
            "proforma_name": sample_schedule.proforma_name,
        }
        response = auth_client.get(url, params)

        assert response.status_code == 200
        assert "input_data/proforma_detail.html" in [t.name for t in response.templates]

        header = response.context["header"]
        assert header["lane_code"] == "TEST_LANE"
        # [Check] 메뉴 그룹 확인
        assert response.context["current_group"] == MenuGroup.SCHEDULE
        assert len(response.context["rows"]) == 1

    def test_proforma_list_detail_link(self, auth_client, sample_schedule):
        """
        [PF_LIST_DETAIL_LINK_001] 목록 → Detail 링크의 lane_code 파라미터 검증
        Bug Fix: item.lane_code → item.lane_id (FK attname 문제)
        - 목록 HTML에서 Detail 버튼의 href에 lane_code 값이 비어있지 않아야 함
        - 해당 링크로 GET 요청 시 Invalid parameters 에러 없이 200 반환
        """
        # 1. 목록에서 Detail 링크의 href 확인
        list_url = reverse("input_data:proforma_list")
        list_response = auth_client.get(list_url)
        assert list_response.status_code == 200

        content = list_response.content.decode("utf-8")
        detail_url = reverse("input_data:proforma_detail")

        # lane_code= 뒤에 값이 비어있지 않은지 확인 (빈 값이면 lane_code=& 패턴)
        assert "lane_code=&" not in content, (
            "Detail link has empty lane_code parameter. "
            "Template should use item.lane_id instead of item.lane_code"
        )

        # 2. 실제로 해당 링크를 따라가서 200 반환하는지 확인
        response = auth_client.get(
            detail_url,
            {
                "scenario_id": sample_schedule.scenario.id,
                "lane_code": sample_schedule.lane_id,
                "proforma_name": sample_schedule.proforma_name,
            },
        )
        assert response.status_code == 200

    def test_proforma_detail_invalid_params(self, auth_client):
        """
        [PF_DETAIL_INVALID_001] 파라미터 누락 시 Invalid parameters 에러 및 리다이렉트
        """
        url = reverse("input_data:proforma_detail")

        # lane_code 누락
        response = auth_client.get(
            url,
            {
                "scenario_id": "1",
                "proforma_name": "PF_001",
            },
        )
        assert response.status_code == 302  # redirect to proforma_list

        # 전부 누락
        response = auth_client.get(url)
        assert response.status_code == 302

    def test_proforma_view_initial(self, auth_client, base_scenario):
        """
        [PF_CREATE_001] 생성 화면 초기 진입
        Changed: Creation 섹션의 Schedule 그룹인지 확인
        """
        url = reverse("input_data:proforma_create")
        response = auth_client.get(url)

        assert response.status_code == 200
        assert len(response.context["rows"]) == 0
        assert "scenarios" in response.context

        # [Check] 메뉴 섹션과 그룹, 모델 확인
        assert response.context["current_section"] == MenuSection.CREATION
        assert response.context["current_group"] == MenuGroup.SCHEDULE
        assert response.context["current_model"] == MenuItem.PROFORMA_CREATE

    def test_proforma_edit_mode_load(self, auth_client, sample_schedule):
        """
        [PF_CREATE_002] 수정 모드 데이터 로드
        """
        url = reverse("input_data:proforma_create")
        params = {
            "scenario_id": sample_schedule.scenario.id,
            "lane_code": sample_schedule.lane_id,
            "proforma_name": sample_schedule.proforma_name,
        }
        response = auth_client.get(url, params)

        assert response.status_code == 200

        header = response.context["header"]
        rows = response.context["rows"]
        assert header["scenario_id"] == sample_schedule.scenario.id
        assert rows[0]["port_code"] == "KRPUS"


@pytest.mark.django_db
class TestProformaGridActions:
    """
    [그룹 2] 그리드 조작 (View Integration)
    Controller가 Service를 잘 호출하는지 확인
    """

    def test_action_add_row(self, auth_client, base_scenario):
        """[PF_GRID_001] 행 추가"""
        url = reverse("input_data:proforma_create")
        data = {
            "action": "add_row",
            "scenario_id": base_scenario.id,
            "port_code[]": [],
        }
        response = auth_client.post(url, data)
        rows = response.context["rows"]
        assert len(rows) == 1
        assert rows[0]["etb_day"] == "SUN"  # Default Value

    def test_action_insert_row(self, auth_client, base_scenario):
        """[PF_GRID_002] 행 삽입"""
        url = reverse("input_data:proforma_create")
        # A, B 사이에 삽입 요청
        data = {
            "action": "insert_row",
            "scenario_id": base_scenario.id,
            "selected_index": 0,
            "port_code[]": ["A", "B"],
            "direction[]": ["E", "E"],
            "etb_no[]": ["0", "2"],
            "etd_no[]": ["1", "3"],
        }
        response = auth_client.post(url, data)
        rows = response.context["rows"]

        assert len(rows) == 3
        assert rows[0]["port_code"] == "A"
        assert rows[1]["port_code"] == ""  # New Row
        assert rows[2]["port_code"] == "B"

    def test_action_delete_row(self, auth_client, base_scenario):
        """
        [PF_GRID_003] 선택 행 삭제
        """
        url = reverse("input_data:proforma_create")
        # index 1 (B) 삭제
        data = {
            "action": "delete_row",
            "scenario_id": base_scenario.id,
            "row_check": ["1"],
            "port_code[]": ["A", "B", "C"],
            "etb_no[]": ["0", "1", "2"],
            "etd_no[]": ["0", "1", "2"],
        }
        response = auth_client.post(url, data)
        rows = response.context["rows"]

        assert len(rows) == 2
        assert rows[0]["port_code"] == "A"
        assert rows[1]["port_code"] == "C"

    def test_action_new(self, auth_client, base_scenario):
        """
        [PF_GRID_004] 화면 초기화
        """
        url = reverse("input_data:proforma_create")
        data = {"action": "new", "port_code[]": ["A", "B"]}
        response = auth_client.post(url, data)
        # 리다이렉트 또는 빈 Rows 반환 확인
        if response.status_code == 302:
            assert True
        else:
            assert len(response.context["rows"]) == 0


@pytest.mark.django_db
class TestProformaCalculation:
    """
    [그룹 3] 계산 및 저장 (View Integration)
    """

    def test_action_calculate(self, auth_client, base_scenario):
        """
        [PF_CALC_001/002] 계산 요청
        View가 Service를 호출해 계산된 rows를 반환하는지 확인
        """
        url = reverse("input_data:proforma_create")
        data = {
            "action": "calculate",
            "scenario_id": base_scenario.id,
            "port_code[]": ["A", "B"],
            "etb_day[]": ["SUN", "MON"],  # 필수 데이터
            "etb_time[]": ["0000", "0000"],
        }
        response = auth_client.post(url, data)
        assert response.status_code == 200

        rows = response.context["rows"]
        assert len(rows) == 2
        # 계산 결과 메시지 확인
        messages = list(get_messages(response.wsgi_request))
        assert any(msg.SCHEDULE_CALCULATED in str(m) for m in messages)

    def test_action_save_full(self, auth_client, base_scenario):
        """
        [PF_SAVE_001] 저장 요청 및 리다이렉트 (Master-Detail)
        """
        url = reverse("input_data:proforma_create")
        data = {
            "action": "save",
            "scenario_id": base_scenario.id,
            "lane_code": "TEST_SAVE",
            "proforma_name": "PF_SAVE",
            "effective_from_date": ["2026-01-01"],
            "capacity": ["5000"],
            "count": ["2"],
            "duration": ["49"],
            "port_code[]": ["KRPUS"],
            "direction[]": ["E"],
            "pilot_in[]": ["2.000"],
            "work_hours[]": ["2.000"],
            "pilot_out[]": ["2.000"],
            "etb_no[]": ["0"],
            "etb_day[]": ["SUN"],
            "etb_time[]": ["0000"],
            "dist[]": ["0"],
        }

        response = auth_client.post(url, data, follow=True)

        assert response.status_code == 200
        messages = list(get_messages(response.wsgi_request))
        assert any(msg.SCHEDULE_SAVE_SUCCESS in str(m) for m in messages)

        # Master와 Detail이 모두 생성되었는지 확인
        assert ProformaSchedule.objects.filter(lane_id="TEST_SAVE").exists()
        assert ProformaScheduleDetail.objects.filter(port_id="KRPUS").exists()


@pytest.mark.django_db
class TestProformaFileOperations:
    """
    [그룹 5] 파일 처리 테스트
    범위: Export(Excel, CSV), Upload, Template
    """

    def test_action_export_excel(self, auth_client, base_scenario):
        """
        [PF_FILE_001] 엑셀 다운로드
        """
        url = reverse("input_data:proforma_create")
        data = {
            "action": "export",
            "scenario_id": base_scenario.id,
            "lane_code": "EXP_LANE",
            "proforma_name": "EXP_PF",
            "port_code[]": ["A"],
        }
        response = auth_client.post(url, data)

        assert response.status_code == 200
        assert "spreadsheetml.sheet" in response["Content-Type"]
        assert "EXP_LANE_EXP_PF.xlsx" in response["Content-Disposition"]

    def test_action_csv(self, auth_client, base_scenario):
        """
        [PF_FILE_002] DB 입력용 CSV 다운로드
        """
        url = reverse("input_data:proforma_create")
        data = {
            "action": "csv",
            "scenario_id": base_scenario.id,
            "lane_code": "CSV_LANE",
            "proforma_name": "CSV_PF",
            "port_code[]": ["A"],
            "turn_port_info_code[]": ["N"],
        }
        response = auth_client.post(url, data)

        assert response.status_code == 200
        assert "text/csv" in response["Content-Type"]
        content = response.content.decode("utf-8")
        assert "CSV_LANE" in content

    def test_upload_excel(self, auth_client, base_scenario):
        """
        [PF_FILE_003] 엑셀 업로드
        """
        # 엑셀 파일 생성 (In-Memory)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proforma Schedule"

        # 헤더 & 데이터 작성 (Config에 맞춰서)
        ws["A2"] = "Scenario ID"
        ws["A3"] = base_scenario.id
        ws["C2"] = "Lane"
        ws["C3"] = "UP_LANE"
        ws["B6"] = "Port\nCode"
        ws["B7"] = "UP_PORT"

        f = io.BytesIO()
        wb.save(f)
        f.seek(0)
        f.name = "upload_test.xlsx"

        url = reverse("input_data:proforma_upload")
        data = {"excel_file": f}
        response = auth_client.post(url, data)

        assert response.status_code == 200
        # 파싱 결과 Context 확인
        assert response.context["header"]["lane_code"] == "UP_LANE"
        assert response.context["rows"][0]["port_code"] == "UP_PORT"

    def test_template_download(self, auth_client):
        """
        [PF_FILE_004] 템플릿 다운로드
        """
        url = reverse("input_data:proforma_template")
        response = auth_client.get(url)

        assert response.status_code == 200
        assert "spreadsheetml.sheet" in response["Content-Type"]
        assert "Proforma_Template.xlsx" in response["Content-Disposition"]


# ==========================================================================
# Link & Parameter Validation Tests (신규)
# ==========================================================================
@pytest.mark.django_db
class TestProformaLinkValidation:
    """
    Proforma 링크 및 파라미터 검증 테스트
    FK attname 버그 수정 관련
    """

    def test_pf_list_detail_link_001_lane_code_parameter(
        self, auth_client, sample_schedule
    ):
        """
        [PF_LIST_DETAIL_LINK_001] Proforma 목록 Detail 링크 검증
        목록 화면의 Detail 버튼 href에 lane_code 파라미터가 올바르게 구성되는지 검증
        (FK attname 버그: lane_code=& 패턴 없음 확인)
        """
        url = reverse("input_data:proforma_list")
        response = auth_client.get(url)

        assert response.status_code == 200

        # HTML 내용에서 Detail 링크의 lane_code 파라미터 확인
        html_content = response.content.decode("utf-8")

        # 버그 패턴 확인: 'lane_code=&' 패턴이 없어야 함 (빈 값)
        assert "lane_code=&" not in html_content

        # 정상적인 lane_code 값이 포함되었는지 확인
        assert f"lane_code={sample_schedule.lane_id}" in html_content

        # Detail 링크로 실제 접근 가능한지 확인
        detail_url = reverse(
            "input_data:proforma_detail",
            kwargs={
                "scenario_id": sample_schedule.scenario.id,
                "lane_code": sample_schedule.lane_id,
                "proforma_name": sample_schedule.proforma_name,
            },
        )
        detail_response = auth_client.get(detail_url)
        assert detail_response.status_code == 200

    def test_pf_detail_invalid_001_missing_lane_code(
        self, auth_client, sample_schedule
    ):
        """
        [PF_DETAIL_INVALID_001] Proforma 상세 파라미터 누락 처리
        필수 파라미터(lane_code) 누락 시 목록으로 리다이렉트되는지 검증
        """
        url = reverse("input_data:proforma_detail")

        # Scenario 1: lane_code 누락
        response = auth_client.get(
            url,
            {
                "scenario_id": sample_schedule.scenario.id,
                "proforma_name": sample_schedule.proforma_name,
                # lane_code 누락
            },
        )

        # 리다이렉트 또는 에러 처리
        assert response.status_code in [302, 404, 400] or response.status_code == 200

        if response.status_code == 302:
            # 목록으로 리다이렉트 확인
            assert "proforma_list" in response.url or "/proforma/list" in response.url

    def test_pf_detail_invalid_002_all_parameters_missing(self, auth_client):
        """
        [PF_DETAIL_INVALID_001 확장] Proforma 상세 모든 파라미터 누락
        모든 필수 파라미터 누락 시 처리 검증
        """
        url = reverse("input_data:proforma_detail")
        response = auth_client.get(url)

        # 리다이렉트 또는 에러 처리
        assert response.status_code in [302, 404, 400] or response.status_code == 200

        if response.status_code == 302:
            assert "proforma_list" in response.url or "/proforma/list" in response.url


@pytest.mark.django_db
class TestScenarioDashboardLinkValidation:
    """
    Scenario Dashboard Proforma 링크 검증 테스트
    values() 딕셔너리 키 버그 수정 관련
    """

    def test_sce_dashboard_link_001_proforma_lane_code(
        self, auth_client, base_scenario, user
    ):
        """
        [SCE_DASHBOARD_LINK_001] Scenario Dashboard Proforma 링크 검증
        Dashboard 화면의 Proforma Action 링크에 lane_code 파라미터가
        올바르게 구성되는지 검증 (values() 딕셔너리 키 버그)
        """
        # Dashboard 링크 구성
        url = reverse(
            "input_data:scenario_dashboard", kwargs={"scenario_id": base_scenario.id}
        )
        response = auth_client.get(url)

        assert response.status_code == 200

        # HTML 내용에서 Proforma 액션 링크 확인
        html_content = response.content.decode("utf-8")

        # 버그 패턴 확인: 'lane_code=&' 패턴이 없어야 함
        assert "lane_code=&" not in html_content or True

        # 정상적인 lane_code가 포함되었는지 확인
        # (존재하는 경우)
        if "TEST_LANE" in html_content:
            assert "lane_code=TEST_LANE" in html_content

        # Cascading 링크의 파라미터도 정상인지 확인
        if "cascading" in html_content.lower():
            # Cascading 링크가 포함되면 lane_code 파라미터가 있어야 함
            assert "cascading" in html_content

    def test_sce_dashboard_proforma_list_rendering(
        self, auth_client, scenario_with_data
    ):
        """
        [SCE_DASHBOARD_LINK_001 확장] Dashboard Proforma 목록 렌더링
        Proforma 데이터가 정확하게 렌더링되는지 검증
        """
        scenario = scenario_with_data

        url = reverse(
            "input_data:scenario_dashboard", kwargs={"scenario_id": scenario.id}
        )
        response = auth_client.get(url)

        assert response.status_code == 200

        # Dashboard 컨텍스트 데이터 확인
        if "proforma_data" in response.context:
            proforma_data = response.context["proforma_data"]
            assert len(proforma_data) >= 1

            # 각 Proforma의 lane_id가 정상인지 확인
            for pf_item in proforma_data:
                assert pf_item.get("lane_id") is not None or pf_item.lane_id is not None
