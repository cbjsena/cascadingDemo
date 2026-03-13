import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import messages as msg
from common.constants import (
    EXCEL_CONSECUTIVE_EMPTY_LIMIT,
    EXCEL_DEFAULT_DATA_START_ROW,
    EXCEL_DEFAULT_GRID_START_ROW,
    EXCEL_DEFAULT_MERGE_WIDTH,
    EXCEL_DEFAULT_TIME,
    EXCEL_KEY_COLUMN_INDEX,
    EXCEL_LABEL_SUMMARY,
    EXCEL_MAX_SCAN_ROWS,
    EXCEL_NO_COLUMN_INDEX,
    EXCEL_SECTION_BASIC,
    EXCEL_SECTION_GRID,
    EXCEL_TEMPLATE_EMPTY_ROWS,
)


class ExcelManager:
    """Config Dictionary 기반 엑셀 생성/파싱 유틸리티"""

    def __init__(self):
        self._init_styles()

    # =================================================================
    # Public API
    # =================================================================

    def create_template(self, config, header_data=None, rows_data=None):
        """
        엑셀 생성 (템플릿 또는 데이터 익스포트).

        :param config: excel_configs에 정의된 설정 딕셔너리
        :param header_data: Basic Info에 채울 데이터 (없으면 config 예시 사용)
        :param rows_data: Grid에 채울 데이터 리스트 (없으면 빈 행 생성)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config.get("sheet_title", "Sheet1")

        self._write_basic_info(ws, config, header_data)
        self._write_grid_headers(ws, config)
        row_count = self._write_grid_data(ws, config, rows_data)
        self._write_summary_if_needed(ws, config, row_count)
        self._adjust_widths(ws, config.get("col_widths", []))

        return self._save_to_bytes(wb)

    def parse_excel(self, file_obj, config):
        """
        엑셀 파일을 파싱하여 (header_data, rows) 튜플 반환.

        :param file_obj: 업로드된 파일 객체
        :param config: excel_configs에 정의된 설정 딕셔너리
        """
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            raise ValueError(msg.INVALID_EXCEL_FILE.format(error=str(e))) from e

        header_data = self._parse_basic_headers(ws, config)
        rows = self._parse_grid_rows(ws, config)
        return header_data, rows

    # =================================================================
    # Template 작성 - 단계별 분리
    # =================================================================

    def _write_basic_info(self, ws, config, header_data):
        """Basic Information 섹션 작성"""
        self._write_section_title(ws, "A1", EXCEL_SECTION_BASIC)
        source = header_data if header_data else config.get("basic_examples", {})

        for item in config["basic_headers"]:
            h_loc, text, key, width = self._unpack_header_item(item)
            val = source.get(key, "")
            self._write_merged_header_value(ws, h_loc, text, val, width)

    def _write_grid_headers(self, ws, config):
        """Grid 헤더 행 작성"""
        grid_start_row = config.get("start_row_grid", EXCEL_DEFAULT_GRID_START_ROW)
        self._write_section_title(ws, f"A{grid_start_row - 1}", EXCEL_SECTION_GRID)

        for text, _, col_idx in config["grid_headers"]:
            cell = ws.cell(row=grid_start_row, column=col_idx, value=text)
            self._apply_header_style(cell)

    def _write_grid_data(self, ws, config, rows_data):
        """Grid 데이터 행 작성. 작성된 행 수를 반환."""
        data_start_row = config.get("start_row_data", EXCEL_DEFAULT_DATA_START_ROW)
        target_rows = rows_data if rows_data else range(EXCEL_TEMPLATE_EMPTY_ROWS)
        grid_examples = config.get("grid_examples", {}) if not rows_data else {}

        for i, row_item in enumerate(target_rows):
            curr_row = data_start_row + i
            self._write_no_cell(ws, curr_row, i + 1)
            self._write_data_cells(
                ws,
                curr_row,
                config["grid_headers"],
                row_item,
                rows_data,
                grid_examples,
                i,
            )

        return len(target_rows)

    def _write_no_cell(self, ws, row, seq):
        """No. 컬럼 (순번) 작성"""
        cell = ws.cell(row=row, column=EXCEL_NO_COLUMN_INDEX, value=seq)
        self._apply_body_style(cell)

    def _write_data_cells(
        self, ws, curr_row, grid_headers, row_item, rows_data, grid_examples, row_index
    ):
        """한 행의 데이터 컬럼들을 작성"""
        for _, key, col_idx in grid_headers:
            if col_idx == EXCEL_NO_COLUMN_INDEX:
                continue

            val = self._resolve_cell_value(
                key, row_item, rows_data, grid_examples, row_index
            )
            cell = ws.cell(row=curr_row, column=col_idx, value=val)
            self._apply_body_style(cell)

    def _write_summary_if_needed(self, ws, config, row_count):
        """Summary Row 작성 (config에 summary_cols가 있을 때만)"""
        if "summary_cols" not in config or row_count == 0:
            return

        data_start_row = config.get("start_row_data", EXCEL_DEFAULT_DATA_START_ROW)
        end_row = data_start_row + row_count - 1
        self._write_summary(
            ws,
            summary_row=end_row + 1,
            data_start=data_start_row,
            data_end=end_row,
            sum_cols=config["summary_cols"],
            total_cols=len(config["grid_headers"]),
            derived_formulas=config.get("derived_formulas", []),
        )

    # =================================================================
    # 파싱 - 단계별 분리
    # =================================================================

    def _parse_basic_headers(self, ws, config):
        """Basic Information 헤더 파싱"""
        header_data = {}
        for item in config["basic_headers"]:
            h_loc, _, key, _ = self._unpack_header_item(item)
            row_idx, col = openpyxl.utils.coordinate_to_tuple(h_loc)
            val_cell = ws.cell(row=row_idx + 1, column=col)
            header_data[key] = self._get_safe_value(val_cell)
        return header_data

    def _parse_grid_rows(self, ws, config):
        """Grid 데이터 행 파싱"""
        rows = []
        start_row = config.get("start_row_data", EXCEL_DEFAULT_DATA_START_ROW)
        empty_count = 0

        for row_idx in range(start_row, start_row + EXCEL_MAX_SCAN_ROWS):
            check_val = ws.cell(row=row_idx, column=EXCEL_KEY_COLUMN_INDEX).value

            if str(check_val).strip() == EXCEL_LABEL_SUMMARY:
                break

            if not check_val or str(check_val).strip() == "":
                empty_count += 1
                if empty_count >= EXCEL_CONSECUTIVE_EMPTY_LIMIT:
                    break
                continue

            empty_count = 0
            row_data = self._parse_single_row(ws, row_idx, config["grid_headers"])
            rows.append(row_data)

        return rows

    def _parse_single_row(self, ws, row_idx, grid_headers):
        """단일 행의 모든 컬럼 파싱"""
        row_data = {}
        for _, key, col_idx in grid_headers:
            val = ws.cell(row=row_idx, column=col_idx).value
            row_data[key] = (
                self._parse_time(val)
                if "time" in key
                else (val if val is not None else "")
            )
        return row_data

    # =================================================================
    # Internal Helpers
    # =================================================================

    @staticmethod
    def _unpack_header_item(item):
        """basic_headers 튜플 (3개 또는 4개 요소)을 통일된 4-tuple로 반환"""
        if len(item) == 4:
            return item
        h_loc, text, key = item
        return h_loc, text, key, EXCEL_DEFAULT_MERGE_WIDTH

    @staticmethod
    def _resolve_cell_value(key, row_item, rows_data, grid_examples, row_index):
        """데이터 행의 셀 값 결정: 실제 데이터 > 첫 행 예시 > 빈 값"""
        if rows_data:
            return row_item.get(key, "")
        if row_index == 0 and grid_examples:
            return grid_examples.get(key, "")
        return ""

    def _write_section_title(self, ws, loc, text):
        cell = ws[loc]
        cell.value = text
        cell.font = Font(bold=True, size=12)

    def _write_merged_header_value(self, ws, h_loc, text, value="", width=3):
        cell = ws[h_loc]
        cell.value = text
        self._apply_header_style(cell)

        end_col_idx = cell.column + width - 1
        ws.merge_cells(
            start_row=cell.row,
            start_column=cell.column,
            end_row=cell.row,
            end_column=end_col_idx,
        )

        v_row = cell.row + 1
        v_cell = ws.cell(row=v_row, column=cell.column)
        v_cell.value = value
        self._apply_body_style(v_cell)
        ws.merge_cells(
            start_row=v_row,
            start_column=v_cell.column,
            end_row=v_row,
            end_column=end_col_idx,
        )

    def _write_summary(
        self,
        ws,
        summary_row,
        data_start,
        data_end,
        sum_cols,
        total_cols,
        derived_formulas=None,
    ):
        """Summary 행 작성 (SUM 수식 + 파생 수식)"""
        label = ws.cell(
            row=summary_row, column=EXCEL_KEY_COLUMN_INDEX, value=EXCEL_LABEL_SUMMARY
        )
        self._apply_summary_style(label)

        # SUM 수식
        for col_idx, col_letter in sum_cols.items():
            formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell = ws.cell(row=summary_row, column=col_idx, value=formula)
            self._apply_summary_style(cell)

        # 파생 수식 (예: Speed = Dist / SeaTime)
        for df in derived_formulas or []:
            target_col = df["target_col"]
            formula_template = df["formula"]
            formula = formula_template.format(row=summary_row)
            cell = ws.cell(row=summary_row, column=target_col, value=formula)
            self._apply_summary_style(cell)

        # 빈 셀에도 summary 스타일 적용
        for c in range(1, total_cols + 1):
            if not ws.cell(row=summary_row, column=c).value:
                self._apply_summary_style(ws.cell(row=summary_row, column=c))

    @staticmethod
    def _save_to_bytes(wb):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _adjust_widths(self, ws, col_widths):
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    # =================================================================
    # Styles
    # =================================================================

    def _init_styles(self):
        self.bold_font = Font(bold=True)
        self.center_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.header_fill = PatternFill(
            start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"
        )
        self.summary_fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )

    def _apply_header_style(self, cell):
        cell.font = self.bold_font
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.fill = self.header_fill

    def _apply_body_style(self, cell):
        cell.alignment = self.center_align
        cell.border = self.thin_border

    def _apply_summary_style(self, cell):
        cell.font = self.bold_font
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.fill = self.summary_fill

    @staticmethod
    def _get_safe_value(cell):
        return cell.value if cell.value is not None else ""

    @staticmethod
    def _parse_time(val):
        if val is None:
            return EXCEL_DEFAULT_TIME
        if hasattr(val, "strftime"):
            return val.strftime("%H%M")
        s = str(val).replace(":", "").replace(".", "").strip()
        return s.zfill(4)[:4] if s.isdigit() else EXCEL_DEFAULT_TIME
